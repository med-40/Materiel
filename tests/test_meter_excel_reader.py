from io import BytesIO

from openpyxl import Workbook

from app.modules.meter_readings.excel_reader import load_meter_workbook


def make_workbook_with_import_sheet_on_second_tab():
    workbook = Workbook()
    first = workbook.active
    first.title = "Sheet1"
    first["A1"] = "صفحة معلومات"

    second = workbook.create_sheet("قراءات")
    second.append([
        "الطراز\u00a0",
        "رقم التسجيل",
        "التاريخ",
        "الكيلومترات",
        "الساعات",
        "حالة العداد",
    ])
    second.append(["طراز-1", "REG-1", "19/08/2026", 1200, None, "يعمل"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def test_load_meter_workbook_selects_sheet_containing_import_headers():
    workbook = load_meter_workbook(make_workbook_with_import_sheet_on_second_tab())
    assert workbook.active.title == "قراءات"

    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][0] == "الطراز"
    assert rows[0][1] == "رقم التسجيل"
    assert rows[0][2] == "التاريخ"
    assert rows[0][3] == "الكيلومترات"
    assert rows[0][5] == "حالة العداد"
    assert rows[1][3] == 1200
