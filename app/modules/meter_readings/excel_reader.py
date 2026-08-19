from __future__ import annotations
from io import BytesIO
from zipfile import ZipFile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def tag(n):
    return f"{{{NS}}}{n}"


def col_index(ref):
    n = 0
    for ch in "".join(c for c in ref if c.isalpha()).upper():
        n = n * 26 + ord(ch) - 64
    return max(0, n - 1)


class RawSheet:
    def __init__(self, rows):
        self.rows = rows

    def iter_rows(self, values_only=True):
        return iter(self.rows)


class RawWorkbook:
    def __init__(self, rows):
        self.active = RawSheet(rows)


def _clean_header_text(value):
    if value is None:
        return ""
    # Excel may preserve RTL/LTR marks and zero-width characters in copied headers.
    text = str(value).replace("\ufeff", "")
    text = "".join(
        ch for ch in text
        if ch not in "\u200e\u200f\u202a\u202b\u202c\u202d\u202e\u2066\u2067\u2068\u2069"
    )
    # Normalize the non-breaking space commonly produced by mobile Excel/apps.
    return text.replace("\u00a0", " ").strip()


def _header_score(value):
    """Score likely meter-import headers without changing the workbook data."""
    text = _clean_header_text(value).lower()
    if not text:
        return 0
    compact = "".join(ch for ch in text if ch.isalnum())
    if compact in {
        "الطراز", "model", "equipmentmodel",
        "رقمالتسجيل", "التسجيل", "registration", "registrationnumber",
        "immatriculation", "matricule", "reg",
        "التاريخ", "تاريخالقراءة", "readingdate", "date",
        "الكيلومترات", "كيلومترات", "الكيلومتر", "كيلومتر", "الكم",
        "عدادالكم", "عدادالكيلومترات", "عدادكم", "الكلم", "كلم", "كم",
        "km", "kilometers", "kilometres", "odometer",
        "الساعات", "ساعات", "ساعة", "عدادالساعات", "hours", "hour",
        "hourmeter", "حالهالعداد", "حالةالعداد", "status", "equipmentstatus",
        "operationalstatus", "نوعالعتاد", "equipmenttype", "type",
    }:
        return 3
    if "طراز" in text or "model" in text:
        return 2
    if "تسجيل" in text or "registration" in text or "matricule" in text or "immatriculation" in text:
        return 2
    if "تاريخ" in text or text.endswith("date"):
        return 2
    if "كيلو" in text or "كلم" in text or "odometer" in text or text.endswith("km"):
        return 2
    if "ساع" in text or "hour" in text:
        return 2
    if "حاله" in text or "حالة" in text or "status" in text:
        return 2
    return 0


def _sheet_score(sheet):
    """Return a score for a worksheet that looks like the import sheet."""
    score = 0
    non_empty = 0
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        row_score = sum(_header_score(value) for value in row)
        score = max(score, row_score)
        non_empty += sum(1 for value in row if value is not None and str(value).strip())
    return score, non_empty


def _select_import_sheet(workbook):
    """Select the worksheet containing the import table instead of assuming active."""
    sheets = list(workbook.worksheets)
    if not sheets:
        return workbook

    ranked = [(_sheet_score(sheet), index, sheet) for index, sheet in enumerate(sheets)]
    # Prefer the strongest header match. If none looks like an import table,
    # retain the old active-sheet behavior, avoiding surprises for other XLSX files.
    best_score, _, best_sheet = max(ranked, key=lambda item: (item[0][0], item[0][1]))
    if best_score[0] > 0:
        workbook.active = best_sheet
    return workbook


def _normalize_import_headers(workbook):
    """Normalize historical/exported headers without changing the user's data."""
    try:
        workbook = _select_import_sheet(workbook)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, max_row=20):
            for cell in row:
                # Only text cells are headers/candidate headers. Never stringify
                # dates or numeric values: doing so breaks Excel date parsing.
                if not isinstance(cell.value, str):
                    continue
                value = _clean_header_text(cell.value)
                if value == "نوع العداد":
                    cell.value = "نوع العتاد"
                elif value == "حاله العداد":
                    cell.value = "حالة العداد"
                elif value:
                    cell.value = value
    except Exception:
        pass
    return workbook


def _normalize_raw_headers(rows):
    for row in rows[:20]:
        for idx, value in enumerate(row):
            # Preserve numeric Excel serials and other non-text data.
            if not isinstance(value, str):
                continue
            value = _clean_header_text(value)
            if value == "نوع العداد":
                value = "نوع العتاد"
            elif value == "حاله العداد":
                value = "حالة العداد"
            row[idx] = value
    return rows


def raw_xlsx(data):
    with ZipFile(BytesIO(data)) as z:
        names = set(z.namelist())
        shared = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root.findall(tag("si")):
                shared.append("".join(t.text or "" for t in si.iter(tag("t"))))
        sheet = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in names and "xl/_rels/workbook.xml.rels" in names:
            wb = ET.fromstring(z.read("xl/workbook.xml"))
            rr = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            rels = {r.attrib.get("Id"): r.attrib.get("Target") for r in rr}
            sheets = wb.find(tag("sheets"))
            first = next(iter(sheets), None) if sheets is not None else None
            if first is not None:
                target = rels.get(first.attrib.get(f"{{{REL}}}id"))
                if target:
                    target = target.lstrip("/")
                    sheet = target if target.startswith("xl/") else "xl/" + target
        root = ET.fromstring(z.read(sheet))
        sd = root.find(tag("sheetData"))
        rows = []
        if sd is None:
            return RawWorkbook([])
        for row in sd.findall(tag("row")):
            cells = {}
            mx = -1
            for c in row.findall(tag("c")):
                idx = col_index(c.attrib.get("r", "A1"))
                mx = max(mx, idx)
                typ = c.attrib.get("t")
                v = c.find(tag("v"))
                raw = v.text if v is not None else None
                value = None
                if typ == "inlineStr":
                    el = c.find(tag("is"))
                    value = "".join(t.text or "" for t in el.iter(tag("t"))) if el is not None else ""
                elif raw is not None and typ == "s":
                    try:
                        value = shared[int(raw)]
                    except (ValueError, IndexError):
                        value = raw
                elif raw is not None and typ == "b":
                    value = raw == "1"
                elif raw is not None:
                    try:
                        value = float(raw)
                        value = int(value) if value.is_integer() else value
                    except (TypeError, ValueError):
                        value = raw
                cells[idx] = value
            rows.append([cells.get(i) for i in range(mx + 1)])
        return RawWorkbook(_normalize_raw_headers(rows))


def load_meter_workbook(stream):
    data = stream.read()
    if hasattr(stream, "seek"):
        stream.seek(0)
    try:
        workbook = load_workbook(BytesIO(data), read_only=False, data_only=True)
        return _normalize_import_headers(workbook)
    except Exception as exc:
        try:
            return raw_xlsx(data)
        except Exception as raw_exc:
            raise ValueError(
                f"تعذر قراءة ملف Excel. تأكد أن الملف بصيغة XLSX سليمة. التفاصيل: {exc}"
            ) from raw_exc
